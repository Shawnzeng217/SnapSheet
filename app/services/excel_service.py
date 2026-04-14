"""
Excel Service / Excel生成服务
将OCR识别结果转换为Excel文件，支持两种模式：
1. data_only: 仅提取数据字段列表
2. table_structure（默认）: 完整还原表格结构（含合并单元格、表头样式）
"""
import os
import re
import json
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models import OCRResult
from app.config import settings

logger = logging.getLogger(__name__)


class ExcelService:
    """Excel生成服务"""

    TEMPLATE_DIR = settings.templates_dir
    OUTPUT_DIR = settings.output_dir

    # 样式常量
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    HEADER_FONT = Font(bold=True, size=10)
    BODY_FONT = Font(size=10)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def __init__(self):
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)

    async def generate(
        self,
        ocr_result: OCRResult,
        template_id: str | None = None,
        output_mode: str = "data_only",
    ) -> str:
        """
        生成Excel文件 / Generate Excel file from OCR result

        逻辑优先级 / Priority:
        1. 有模板 → 以模板为准，智能填入手写内容
        2. 无模板 + 有表格HTML → 自动还原表格结构
        3. 无模板 + 无表格 → 字段列表
        """
        if template_id:
            logger.info(f"generate: template_id={template_id!r} -> template_restore")
            return await self._generate_template_restore(ocr_result, template_id)
        elif ocr_result.table_html:
            logger.info(f"generate: no template, has table_html -> table_structure")
            return await self._generate_table_structure(ocr_result)
        else:
            return await self._generate_data_only(ocr_result)

    async def generate_raw(self, ocr_result: OCRResult) -> str:
        """
        生成原始识别结果Excel（不使用模板）
        Generate raw OCR result Excel without template
        """
        if ocr_result.table_html:
            return await self._generate_table_structure(ocr_result)
        else:
            return await self._generate_data_only(ocr_result)

    async def _generate_table_structure(self, ocr_result: OCRResult) -> str:
        """
        表格结构还原模式 - 从HTML表格完整还原到Excel（含合并单元格）
        Table structure mode - fully restore HTML table to Excel with merged cells
        """
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认sheet

        for idx, html in enumerate(ocr_result.table_html):
            sheet_name = f"Table {idx + 1}" if len(ocr_result.table_html) > 1 else "Sheet1"
            ws = wb.create_sheet(title=sheet_name)
            self._html_table_to_sheet(ws, html)

        output_path = os.path.join(self.OUTPUT_DIR, f"{ocr_result.task_id}_table.xlsx")
        wb.save(output_path)
        return output_path

    def _html_table_to_sheet(self, ws, html: str) -> None:
        """
        将HTML表格写入Excel worksheet，正确处理 colspan/rowspan 合并单元格
        Convert HTML table to Excel worksheet with proper colspan/rowspan merging
        """
        # 解析所有行和单元格
        cells_data = self._parse_html_cells(html)
        if not cells_data:
            return

        # 构建网格，处理 colspan/rowspan 占位
        # grid[row][col] = (text, is_header, is_origin)
        # occupied[row][col] = True 表示被前面的合并单元格占据
        grid: dict[int, dict[int, tuple[str, bool]]] = {}
        merges: list[tuple[int, int, int, int]] = []  # (r1, c1, r2, c2) 1-indexed
        occupied: dict[tuple[int, int], bool] = {}

        excel_row = 0  # 当前写入的Excel行 (0-indexed)

        for row_cells, is_thead in cells_data:
            # 找到当前行里第一个未被占据的列
            col = 0
            cell_idx = 0

            for text, is_th, colspan, rowspan in row_cells:
                # 跳过被占据的列
                while occupied.get((excel_row, col)):
                    col += 1

                is_header = is_thead or is_th

                # 写入当前单元格
                grid.setdefault(excel_row, {})[col] = (text, is_header)

                # 处理合并
                if colspan > 1 or rowspan > 1:
                    # 记录合并区域 (1-indexed for openpyxl)
                    r1 = excel_row + 1
                    c1 = col + 1
                    r2 = excel_row + rowspan
                    c2 = col + colspan
                    merges.append((r1, c1, r2, c2))

                    # 标记占据的区域
                    for dr in range(rowspan):
                        for dc in range(colspan):
                            if dr == 0 and dc == 0:
                                continue
                            occupied[(excel_row + dr, col + dc)] = True

                col += colspan
                cell_idx += 1

            excel_row += 1

        # 写入Excel
        max_col = 0
        for r, cols in grid.items():
            for c, (text, is_header) in cols.items():
                cell = ws.cell(row=r + 1, column=c + 1, value=text)
                cell.border = self.THIN_BORDER
                cell.alignment = self.CENTER_ALIGN if is_header else self.LEFT_ALIGN
                if is_header:
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                else:
                    cell.font = self.BODY_FONT
                max_col = max(max_col, c + 1)

        # 应用合并单元格
        for r1, c1, r2, c2 in merges:
            ws.merge_cells(
                start_row=r1, start_column=c1,
                end_row=r2, end_column=c2,
            )
            # 给合并区域内所有单元格补上边框
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    ws.cell(row=r, column=c).border = self.THIN_BORDER

        # 自适应列宽
        for col_idx in range(1, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        # 中文字符按2计算宽度
                        val = str(cell.value)
                        w = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_len = max(max_len, w)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    def _parse_html_cells(self, html: str) -> list[tuple[list[tuple[str, bool, int, int]], bool]]:
        """
        解析HTML表格为结构化单元格数据
        Returns: [(row_cells, is_thead), ...]
        row_cells: [(text, is_th, colspan, rowspan), ...]
        """
        result = []
        tag_strip = re.compile(r"<[^>]+>")

        # 分析 thead / tbody 区域
        # 先提取 thead 中的行
        thead_match = re.search(r"<thead[^>]*>(.*?)</thead>", html, re.DOTALL | re.IGNORECASE)
        tbody_match = re.search(r"<tbody[^>]*>(.*?)</tbody>", html, re.DOTALL | re.IGNORECASE)

        sections = []
        if thead_match:
            sections.append((thead_match.group(1), True))
        if tbody_match:
            sections.append((tbody_match.group(1), False))

        # 如果没有 thead/tbody，直接解析整个 table
        if not sections:
            table_match = re.search(r"<table[^>]*>(.*)</table>", html, re.DOTALL | re.IGNORECASE)
            if table_match:
                sections.append((table_match.group(1), False))
            else:
                sections.append((html, False))

        tr_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL | re.IGNORECASE)
        cell_pattern = re.compile(
            r"<(t[dh])([^>]*)>(.*?)</\1>",
            re.DOTALL | re.IGNORECASE,
        )

        for section_html, is_thead in sections:
            for tr_match in tr_pattern.finditer(section_html):
                row_html = tr_match.group(1)
                row_cells = []

                for cell_match in cell_pattern.finditer(row_html):
                    tag_name = cell_match.group(1).lower()
                    attrs = cell_match.group(2)
                    inner = cell_match.group(3)

                    is_th = tag_name == "th"

                    # 提取 colspan / rowspan
                    colspan = 1
                    rowspan = 1
                    cs = re.search(r'colspan\s*=\s*["\']?(\d+)', attrs, re.IGNORECASE)
                    rs = re.search(r'rowspan\s*=\s*["\']?(\d+)', attrs, re.IGNORECASE)
                    if cs:
                        colspan = int(cs.group(1))
                    if rs:
                        rowspan = int(rs.group(1))

                    # 清理HTML标签，保留文本（<br>转换为换行符）
                    text = re.sub(r"<br\s*/?>", "\n", inner, flags=re.IGNORECASE)
                    text = tag_strip.sub("", text).strip()

                    row_cells.append((text, is_th, colspan, rowspan))

                if row_cells:
                    result.append((row_cells, is_thead))

        return result

    async def _generate_data_only(self, ocr_result: OCRResult) -> str:
        """
        数据提取模式 - 生成简洁的数据表格
        Data-only mode - generate a clean data table
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Result"

        # 表头 / Header
        headers = ["字段 / Field", "值 / Value", "置信度 / Confidence"]
        header_font = Font(bold=True, size=11)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 数据行 / Data rows
        for i, item in enumerate(ocr_result.items, 2):
            ws.cell(row=i, column=1, value=item.field_name)
            ws.cell(row=i, column=2, value=item.value)
            ws.cell(row=i, column=3, value=round(item.confidence, 3))

        # 调整列宽 / Adjust column width
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15

        output_path = os.path.join(self.OUTPUT_DIR, f"{ocr_result.task_id}_data.xlsx")
        wb.save(output_path)
        return output_path

    async def _generate_template_restore(self, ocr_result: OCRResult, template_id: str) -> str:
        """
        模板还原模式 - 基于原始Excel模板，用OCR识别的手写内容智能填入空白单元格
        Template restore mode - load original Excel template, intelligently fill
        handwritten content into empty cells using grid-position alignment.

        策略 / Strategy (网格对齐):
        1. 将OCR HTML表格解析为二维网格 (0-indexed)
        2. 将模板解析为二维网格 (1-indexed)
        3. 用共同文本作为锚点，计算两个网格的行/列偏移量
        4. 对模板中每个空白单元格，按偏移量在OCR网格中查找对应值
        """
        template_path = self._find_template(template_id)
        if not template_path:
            if ocr_result.table_html:
                return await self._generate_table_structure(ocr_result)
            return await self._generate_data_only(ocr_result)

        wb = load_workbook(template_path)
        ws = wb.active

        # 从OCR结果构建二维网格
        ocr_grid = self._build_ocr_grid(ocr_result)

        if ocr_grid:
            self._grid_align_fill(ws, ocr_grid)
        else:
            # 回退：按顺序填入空白单元格
            self._sequential_fill_template(ws, ocr_result)

        output_path = os.path.join(self.OUTPUT_DIR, f"{ocr_result.task_id}_restored.xlsx")
        wb.save(output_path)
        return output_path

    def _build_ocr_grid(self, ocr_result: OCRResult) -> dict[int, dict[int, str]]:
        """
        从OCR HTML表格构建完整的二维网格 (0-indexed)
        Build a 2D grid from OCR HTML tables (0-indexed)
        多个表格按顺序拼接，行号连续递增
        Multiple tables are concatenated sequentially with continuous row numbers
        """
        if not ocr_result.table_html:
            return {}

        grid: dict[int, dict[int, str]] = {}
        occupied: dict[tuple[int, int], bool] = {}
        excel_row = 0

        for html in ocr_result.table_html:
            cells_data = self._parse_html_cells(html)
            if not cells_data:
                continue

            for row_cells, is_thead in cells_data:
                col = 0
                for text, is_th, colspan, rowspan in row_cells:
                    while occupied.get((excel_row, col)):
                        col += 1
                    grid.setdefault(excel_row, {})[col] = text
                    for dr in range(rowspan):
                        for dc in range(colspan):
                            if dr == 0 and dc == 0:
                                continue
                            occupied[(excel_row + dr, col + dc)] = True
                            grid.setdefault(excel_row + dr, {})[col + dc] = ""
                    col += colspan
                excel_row += 1

        return grid

    def _grid_align_fill(self, ws, ocr_grid: dict[int, dict[int, str]]) -> None:
        """
        通用网格对齐填充 - 适配任意模板
        General-purpose grid-aligned fill - works with any template

        策略 / Strategy:
        1. 用两个网格的共同文本(锚点)建立逐行映射: 模板行 → OCR行
        2. 未映射的行通过邻近行插值
        3. 按行映射+列偏移填入值
        4. 未填的单元格用标签邻接取值(左侧/上方)作为补充
        """
        from collections import Counter, defaultdict

        # ── 1. 构建模板网格视图（处理合并单元格）──
        merged_map: dict[tuple[int, int], str] = {}
        for mr in ws.merged_cells.ranges:
            val = ws.cell(row=mr.min_row, column=mr.min_col).value
            if val:
                val_str = str(val).strip()
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        merged_map[(r, c)] = val_str

        def get_cell_text(row: int, col: int) -> str:
            if (row, col) in merged_map:
                return merged_map[(row, col)]
            v = ws.cell(row=row, column=col).value
            return str(v).strip() if v is not None else ""

        max_tpl_row = ws.max_row or 1
        max_tpl_col = ws.max_column or 1

        # ── 2. 构建文本 → 位置索引 ──
        # 对于合并单元格，只用左上角位置，避免同一文本产生多个位置导致唯一性判断失败
        merged_topleft: set[tuple[int, int]] = set()
        merged_slaves_set: set[tuple[int, int]] = set()
        for mr in ws.merged_cells.ranges:
            merged_topleft.add((mr.min_row, mr.min_col))
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if not (r == mr.min_row and c == mr.min_col):
                        merged_slaves_set.add((r, c))

        tpl_text_positions: dict[str, list[tuple[int, int]]] = {}
        for r in range(1, max_tpl_row + 1):
            for c in range(1, max_tpl_col + 1):
                # 跳过合并单元格的从属位置，只保留左上角
                if (r, c) in merged_slaves_set:
                    continue
                text = get_cell_text(r, c)
                if text:
                    norm = self._normalize(text)
                    if norm:
                        tpl_text_positions.setdefault(norm, []).append((r, c))

        ocr_text_positions: dict[str, list[tuple[int, int]]] = {}
        for r, cols in ocr_grid.items():
            for c, text in cols.items():
                if text and text.strip():
                    norm = self._normalize(text.strip())
                    if norm:
                        ocr_text_positions.setdefault(norm, []).append((r, c))

        # ── 3. 找锚点：两侧都出现的文本 ──
        # anchor_texts: 出现在两个网格中的结构性文本(标签/表头)，用于过滤
        anchor_pairs: list[tuple[int, int, int, int]] = []  # (tpl_r, tpl_c, ocr_r, ocr_c)
        anchor_texts: set[str] = set()

        # 第1层：两侧都唯一出现的文本（最可靠）
        for norm_text, tpl_pos in tpl_text_positions.items():
            ocr_pos = ocr_text_positions.get(norm_text)
            if not ocr_pos:
                continue
            if len(tpl_pos) == 1 and len(ocr_pos) == 1:
                anchor_pairs.append((tpl_pos[0][0], tpl_pos[0][1], ocr_pos[0][0], ocr_pos[0][1]))
                anchor_texts.add(norm_text)

        # 第2层：非唯一文本，按顺序配对补充（始终执行，而非仅作为后备）
        # 对称出现的重复标签（如两段表头中的 "STARTING NUMBER"）用序数配对
        for norm_text, tpl_pos in tpl_text_positions.items():
            if norm_text in anchor_texts:
                continue  # 已作为唯一锚点
            ocr_pos = ocr_text_positions.get(norm_text)
            if not ocr_pos:
                continue
            sorted_tpl = sorted(tpl_pos)
            sorted_ocr = sorted(ocr_pos)
            # 仅配对数量相等的情况以保证可靠性
            if len(sorted_tpl) == len(sorted_ocr):
                for (tr, tc), (or_, oc) in zip(sorted_tpl, sorted_ocr):
                    anchor_pairs.append((tr, tc, or_, oc))
                anchor_texts.add(norm_text)

        if not anchor_pairs:
            # 第3层：子串匹配 - OCR常将多行标签合并为一个单元格
            for tpl_norm, tpl_pos in tpl_text_positions.items():
                if len(tpl_norm) < 2:
                    continue
                # 收集所有匹配的OCR位置（可能来自多个OCR归一化文本）
                matching_ocr_positions: list[tuple[int, int]] = []
                for ocr_norm, ocr_pos in ocr_text_positions.items():
                    if tpl_norm in ocr_norm or ocr_norm in tpl_norm:
                        matching_ocr_positions.extend(ocr_pos)
                        anchor_texts.add(ocr_norm)
                if matching_ocr_positions:
                    # 按顺序配对（第1个模板→第1个OCR，第2个→第2个）
                    sorted_tpl = sorted(tpl_pos)
                    sorted_ocr = sorted(matching_ocr_positions)
                    for i, (tr, tc) in enumerate(sorted_tpl):
                        if i < len(sorted_ocr):
                            or_, oc = sorted_ocr[i]
                        else:
                            or_, oc = sorted_ocr[-1]
                        anchor_pairs.append((tr, tc, or_, oc))
                    anchor_texts.add(tpl_norm)

        if not anchor_pairs:
            return

        # ── 4. 建立行映射: 模板行 → OCR行 ──
        # 逐行投票：每个模板行对应哪个OCR行
        row_votes: dict[int, Counter[int]] = defaultdict(Counter)
        col_offset_votes: Counter[int] = Counter()

        for tr, tc, or_, oc in anchor_pairs:
            row_votes[tr][or_] += 1
            col_offset_votes[tc - oc] += 1

        # 直接映射（有锚点的行）
        row_map: dict[int, int] = {}
        for tpl_r, votes in row_votes.items():
            row_map[tpl_r] = votes.most_common(1)[0][0]

        # 全局列偏移
        global_col_off = col_offset_votes.most_common(1)[0][0]

        # 全局行偏移（用于插值兜底）
        row_off_votes: Counter[int] = Counter()
        for tr, or_ in row_map.items():
            row_off_votes[tr - or_] += 1
        global_row_off = row_off_votes.most_common(1)[0][0] if row_off_votes else 0

        # 已排序的有映射的模板行号（用于插值查找邻居）
        mapped_rows = sorted(row_map.keys())

        def get_ocr_row(tpl_r: int) -> int:
            """从最近的锚点行插值推算OCR行号"""
            if tpl_r in row_map:
                return row_map[tpl_r]
            # 找上方最近的已映射行
            above = [mr for mr in mapped_rows if mr < tpl_r]
            below = [mr for mr in mapped_rows if mr > tpl_r]
            if above:
                ref = above[-1]
                return row_map[ref] + (tpl_r - ref)
            if below:
                ref = below[0]
                return row_map[ref] + (tpl_r - ref)
            return tpl_r - global_row_off

        # ── 5. 计算OCR网格的有效边界 ──
        ocr_max_row = max(ocr_grid.keys()) if ocr_grid else 0
        ocr_max_col = 0
        for cols in ocr_grid.values():
            if cols:
                ocr_max_col = max(ocr_max_col, max(cols.keys()))

        valid_tpl_col_min = global_col_off
        valid_tpl_col_max = ocr_max_col + global_col_off

        # ── 5b. 识别模板中的"锚点区段"，防止数据溢出到不相关区域 ──
        # 将连续的锚点行(间距≤3)归为同一区段，然后向下扩展到下一区段前
        # 这样数据行（位于标签行之间）也被纳入区段范围
        raw_sections: list[tuple[int, int]] = []  # (start_row, end_row)
        if mapped_rows:
            sec_start = mapped_rows[0]
            sec_end = mapped_rows[0]
            for mr in mapped_rows[1:]:
                if mr - sec_end <= 3:
                    sec_end = mr
                else:
                    raw_sections.append((sec_start, sec_end))
                    sec_start = mr
                    sec_end = mr
            raw_sections.append((sec_start, sec_end))

        # 扩展每个区段：向下延伸到下一个区段起始行前一行（覆盖数据行）
        # 最后一个区段延伸到OCR有效行范围对应的模板行
        anchor_sections: list[tuple[int, int]] = []
        ocr_max_tpl_row = ocr_max_row + global_row_off  # OCR最大行对应的模板行
        for i, (s, e) in enumerate(raw_sections):
            if i + 1 < len(raw_sections):
                next_start = raw_sections[i + 1][0]
                expanded_end = next_start - 1
            else:
                expanded_end = max(e, ocr_max_tpl_row)
            anchor_sections.append((s, expanded_end))

        def row_in_section(tpl_r: int) -> bool:
            """检查模板行是否在某个锚点区段内"""
            for sec_start, sec_end in anchor_sections:
                if sec_start <= tpl_r <= sec_end:
                    return True
            return False

        # ── 6. 合并单元格从属位置 ──
        merged_slaves: set[tuple[int, int]] = set()
        for mr in ws.merged_cells.ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if not (r == mr.min_row and c == mr.min_col):
                        merged_slaves.add((r, c))

        # ── 7. 按位置映射填入值 ──
        unfilled: list[tuple[int, int]] = []
        filled_cells: set[tuple[int, int]] = set()  # 记录step 7填入的模板位置
        used_ocr_cells: set[tuple[int, int]] = set()  # 记录step 7已消费的OCR位置

        for r in range(1, max_tpl_row + 1):
            for c in range(1, max_tpl_col + 1):
                if (r, c) in merged_slaves:
                    continue

                tpl_text = get_cell_text(r, c)

                # 跳过超出OCR有效列范围的单元格
                if c < valid_tpl_col_min or c > valid_tpl_col_max:
                    continue

                # 跳过不在任何锚点区段内的行（防止填入区段间空白行）
                if not row_in_section(r):
                    continue

                ocr_r = get_ocr_row(r)
                ocr_c = c - global_col_off
                ocr_val = ocr_grid.get(ocr_r, {}).get(ocr_c, "")

                # 如果模板单元格有文本(标签)，检查OCR是否在同一位置有"标签+数据"
                # 例如模板 "HOD:" 对应 OCR "HOD: Jenny" → 更新为 "HOD: Jenny"
                # 仅当OCR原文以模板标签开头时才触发
                if tpl_text and ocr_val and ocr_val.strip():
                    raw_ocr = ocr_val.strip()
                    tpl_clean = tpl_text.strip()
                    if (raw_ocr.startswith(tpl_clean) and
                            len(raw_ocr) > len(tpl_clean) and
                            raw_ocr != tpl_clean):
                        data_part = raw_ocr[len(tpl_clean):].strip()
                        data_part = data_part.lstrip(":： \n")
                        # 排除双语合并标签: 如果data_part本身是模板中的标签就跳过
                        if data_part:
                            data_norm = self._normalize(data_part)
                            is_label = data_norm in tpl_text_positions
                            if not is_label:
                                ws.cell(row=r, column=c).value = f"{tpl_clean} {data_part}"
                                filled_cells.add((r, c))
                                used_ocr_cells.add((ocr_r, ocr_c))
                    continue  # 有文本的单元格不再继续下面的空白填充逻辑

                if tpl_text:
                    continue

                filled = False
                if ocr_val and ocr_val.strip():
                    norm_val = self._normalize(ocr_val.strip())
                    # 只过滤已确认的结构性文本(锚点)，不过滤所有模板文本
                    # 这样 "OK"、数字等数据值即使碰巧与模板某处相同也能被填入
                    if norm_val not in anchor_texts:
                        ws.cell(row=r, column=c).value = ocr_val.strip()
                        filled = True
                        filled_cells.add((r, c))
                        used_ocr_cells.add((ocr_r, ocr_c))

                if not filled:
                    unfilled.append((r, c))

        # ── 8. 补充：标签邻接匹配（处理列偏移不一致的区域）──
        # 对未填充的空白单元格，找模板中相邻的 **原始** 标签文本，
        # 然后在OCR网格中找到该标签，取其邻接单元格的值
        # 注意：只用模板原始文本作为标签，不用step 7填入的数据值
        ocr_norm_to_pos: dict[str, list[tuple[int, int]]] = {}
        for r, cols in ocr_grid.items():
            for c, text in cols.items():
                if text and text.strip():
                    norm = self._normalize(text.strip())
                    if norm:
                        ocr_norm_to_pos.setdefault(norm, []).append((r, c))

        def find_ocr_label_positions(norm_label: str) -> list[tuple[int, int]]:
            """在OCR网格中查找标签位置，支持子串匹配"""
            if norm_label in ocr_norm_to_pos:
                return ocr_norm_to_pos[norm_label]
            # 子串匹配: OCR单元格可能合并了多行标签
            for ocr_norm, positions in ocr_norm_to_pos.items():
                if len(norm_label) >= 2 and (norm_label in ocr_norm or ocr_norm in norm_label):
                    return positions
            return []

        def get_original_text(row: int, col: int) -> str:
            """获取模板原始文本（忽略step 7填入的值）"""
            if (row, col) in filled_cells:
                return ""
            return get_cell_text(row, col)

        for r, c in unfilled:
            if ws.cell(row=r, column=c).value is not None:
                continue

            found = False
            expected_ocr_row = get_ocr_row(r)

            # 策略A: 左侧标签 → OCR中该标签右侧的值
            # 适用于 "标签 | 值 | 标签 | 值" 的水平布局
            # 约束: OCR值列映射回模板列应接近目标列, 且OCR值未被step 7使用
            for cc in range(c - 1, 0, -1):
                label = get_original_text(r, cc)
                if label:
                    norm_label = self._normalize(label)
                    ocr_positions = find_ocr_label_positions(norm_label) if norm_label else []
                    if ocr_positions:
                        nearby = [(olr, olc) for olr, olc in ocr_positions
                                  if abs(olr - expected_ocr_row) <= 2]
                        for olr, olc in nearby:
                            val_pos = (olr, olc + 1)
                            if val_pos in used_ocr_cells:
                                continue
                            # 列验证: OCR值列映射回的模板列应接近目标列c
                            mapped_tpl_col = olc + 1 + global_col_off
                            if abs(mapped_tpl_col - c) > 1:
                                continue
                            rv = ocr_grid.get(olr, {}).get(olc + 1, "")
                            if rv and rv.strip():
                                nrv = self._normalize(rv.strip())
                                if nrv not in anchor_texts:
                                    ws.cell(row=r, column=c).value = rv.strip()
                                    used_ocr_cells.add(val_pos)
                                    found = True
                                    break
                    break  # 只看最近的标签

            if found:
                continue

            # 策略B: 上方标签 → OCR中该标签下方的值
            # 适用于 "标签在上、值在下" 的垂直布局
            # 约束: OCR值行应接近目标行的期望OCR行, 且OCR值未被step 7使用
            for rr in range(r - 1, 0, -1):
                label = get_original_text(rr, c)
                if label:
                    norm_label = self._normalize(label)
                    ocr_positions = find_ocr_label_positions(norm_label) if norm_label else []
                    if ocr_positions:
                        expected_label_ocr_row = get_ocr_row(rr)
                        nearby = [(olr, olc) for olr, olc in ocr_positions
                                  if abs(olr - expected_label_ocr_row) <= 2]
                        for olr, olc in nearby:
                            val_pos = (olr + 1, olc)
                            if val_pos in used_ocr_cells:
                                continue
                            # 行验证: OCR值行应接近当前模板行的期望OCR行
                            if abs(val_pos[0] - expected_ocr_row) > 1:
                                continue
                            bv = ocr_grid.get(olr + 1, {}).get(olc, "")
                            if bv and bv.strip():
                                nbv = self._normalize(bv.strip())
                                if nbv not in anchor_texts:
                                    ws.cell(row=r, column=c).value = bv.strip()
                                    used_ocr_cells.add(val_pos)
                                    found = True
                                    break
                    break

    @staticmethod
    def _normalize(text: str) -> str:
        """
        标准化文本用于匹配：去除空格、标点、统一大小写
        Normalize text for matching: strip whitespace, punctuation, lowercase
        """
        if not text:
            return ""
        # 去除常见标点和空白
        text = re.sub(r"[\s\-_/\\()（）【】\[\]：:，,。.、·]+", "", text)
        return text.lower().strip()

    def _sequential_fill_template(self, ws, ocr_result: OCRResult) -> None:
        """回退：按顺序填入空白单元格 / Fallback: fill empty cells sequentially"""
        empty_cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    empty_cells.append(cell)
        for i, cell in enumerate(empty_cells):
            if i < len(ocr_result.items):
                cell.value = ocr_result.items[i].value

    def _find_template(self, template_id: str) -> str | None:
        """查找模板文件路径"""
        meta_path = os.path.join(self.TEMPLATE_DIR, "meta.json")
        if not os.path.exists(meta_path):
            return None

        with open(meta_path, "r", encoding="utf-8") as f:
            templates = json.load(f)

        for t in templates:
            if t.get("template_id") == template_id:
                filename = t.get("file", "")
                # 兼容旧格式：若存的是带目录的路径，取文件名部分
                filename = os.path.basename(filename.replace("\\", "/"))
                path = os.path.join(self.TEMPLATE_DIR, filename)
                if os.path.exists(path):
                    # 兜底：如果文件是.xls，运行时自动转换
                    if path.lower().endswith(".xls"):
                        xlsx_path = self.convert_xls_to_xlsx(path)
                        t["file"] = os.path.basename(xlsx_path)
                        with open(meta_path, "w", encoding="utf-8") as fw:
                            json.dump(templates, fw, ensure_ascii=False, indent=2)
                        return xlsx_path
                    return path
        return None

    @staticmethod
    def convert_xls_to_xlsx(xls_path: str) -> str:
        """
        将 .xls 文件转换为 .xlsx（保留值、合并单元格）
        Convert legacy .xls to .xlsx preserving values and merged cells
        """
        import xlrd

        xls_wb = xlrd.open_workbook(xls_path, formatting_info=True)
        xls_ws = xls_wb.sheet_by_index(0)

        xlsx_wb = Workbook()
        xlsx_ws = xlsx_wb.active
        xlsx_ws.title = xls_ws.name or "Sheet1"

        # 复制单元格值
        for r in range(xls_ws.nrows):
            for c in range(xls_ws.ncols):
                val = xls_ws.cell_value(r, c)
                cell_type = xls_ws.cell_type(r, c)
                if cell_type == xlrd.XL_CELL_EMPTY:
                    continue
                xlsx_ws.cell(row=r + 1, column=c + 1, value=val)

        # 复制合并单元格
        for rlo, rhi, clo, chi in xls_ws.merged_cells:
            xlsx_ws.merge_cells(
                start_row=rlo + 1, start_column=clo + 1,
                end_row=rhi, end_column=chi,
            )

        xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"
        xlsx_wb.save(xlsx_path)
        return xlsx_path
