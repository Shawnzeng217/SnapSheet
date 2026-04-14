"""
模板管理API / Template Management API
"""
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse
import os
import shutil
import json

import openpyxl

from app.models import APIResponse, TemplateInfo
from app.services.excel_service import ExcelService
from app.config import settings

router = APIRouter()

TEMPLATE_DIR = settings.templates_dir


@router.get("/", response_model=APIResponse)
async def list_templates():
    """列出所有模板 / List all templates"""
    templates = []
    meta_path = os.path.join(TEMPLATE_DIR, "meta.json")
    if os.path.exists(meta_path):
        with open(meta_path, "r", encoding="utf-8") as f:
            templates = json.load(f)
    return APIResponse(data=templates)


@router.post("/upload", response_model=APIResponse)
async def upload_template(
    file: UploadFile = File(..., description="Excel模板文件 / Excel template file"),
    name: str = Form(..., description="模板名称"),
    description: str = Form("", description="模板描述"),
):
    """
    上传Excel模板文件，用于OCR模板匹配
    Upload Excel template for OCR template matching
    """
    os.makedirs(TEMPLATE_DIR, exist_ok=True)

    template_id = name.replace(" ", "_").lower()
    file_ext = os.path.splitext(file.filename or "template.xlsx")[1].lower()

    # 先保存原始文件
    raw_path = os.path.join(TEMPLATE_DIR, f"{template_id}{file_ext}")
    with open(raw_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # .xls → .xlsx 自动转换 / Auto-convert legacy .xls to .xlsx
    if file_ext == ".xls":
        save_path = ExcelService.convert_xls_to_xlsx(raw_path)
        os.remove(raw_path)  # 删除原始.xls
    else:
        save_path = raw_path

    # 更新元数据 / Update metadata
    # 只存文件名，运行时用 TEMPLATE_DIR 拼接完整路径（跨平台兼容）
    meta_path = os.path.join(TEMPLATE_DIR, "meta.json")
    templates = []
    if os.path.exists(meta_path):
        with open(meta_path, "r", encoding="utf-8") as f:
            templates = json.load(f)

    # Remove existing entry with same id
    templates = [t for t in templates if t["template_id"] != template_id]
    templates.append({
        "template_id": template_id,
        "name": name,
        "description": description,
        "file": os.path.basename(save_path),
    })

    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)

    return APIResponse(data={"template_id": template_id, "file": save_path})


@router.delete("/{template_id}", response_model=APIResponse)
async def delete_template(template_id: str):
    """删除模板 / Delete template"""
    meta_path = os.path.join(TEMPLATE_DIR, "meta.json")
    if not os.path.exists(meta_path):
        raise HTTPException(status_code=404, detail="Template not found")

    with open(meta_path, "r", encoding="utf-8") as f:
        templates = json.load(f)

    target = None
    for t in templates:
        if t["template_id"] == template_id:
            target = t
            break

    if not target:
        raise HTTPException(status_code=404, detail="Template not found")

    # Remove file
    file_path = os.path.join(TEMPLATE_DIR, target.get("file", ""))
    if os.path.exists(file_path):
        os.remove(file_path)

    templates = [t for t in templates if t["template_id"] != template_id]
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)

    return APIResponse(message=f"Template {template_id} deleted")


@router.get("/{template_id}/preview", response_model=APIResponse)
async def preview_template(template_id: str, max_rows: int = 50):
    """预览模板内容 / Preview template as HTML table"""
    meta_path = os.path.join(TEMPLATE_DIR, "meta.json")
    if not os.path.exists(meta_path):
        raise HTTPException(status_code=404, detail="Template not found")

    with open(meta_path, "r", encoding="utf-8") as f:
        templates = json.load(f)

    target = None
    for t in templates:
        if t["template_id"] == template_id:
            target = t
            break

    file_path = os.path.join(TEMPLATE_DIR, target.get("file", "")) if target else ""
    if not target or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Template not found")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 收集合并单元格信息 / Collect merged cell ranges
    merges: dict[tuple[int, int], tuple[int, int]] = {}
    skip: set[tuple[int, int]] = set()
    for mg in ws.merged_cells.ranges:
        r1, c1, r2, c2 = mg.min_row, mg.min_col, mg.max_row, mg.max_col
        merges[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    skip.add((r, c))

    html = '<table class="ocr-table-preview">'
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_rows)):
        html += "<tr>"
        for cell in row:
            r, c = cell.row, cell.column
            if (r, c) in skip:
                continue
            val = str(cell.value) if cell.value is not None else ""
            attrs = ""
            if (r, c) in merges:
                rs, cs = merges[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            html += f"<td{attrs}>{val}</td>"
        html += "</tr>"
        row_count += 1
    html += "</table>"

    wb.close()

    return APIResponse(data={
        "template_id": template_id,
        "name": target["name"],
        "html": html,
        "row_count": row_count,
        "truncated": (ws.max_row or 0) > max_rows,
    })
